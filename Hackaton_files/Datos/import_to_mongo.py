"""
Damm Smart Truck · Importació de dades reals a MongoDB Atlas
=============================================================
Llegeix els Excels originals de Damm i els importa a MongoDB.

INSTAL·LACIÓ:
    pip install pymongo openpyxl

ÚS:
    1. Substitueix MONGO_URI per la teva contrasenya real
    2. Substitueix EXCEL_DIR pel camí als teus Excels
    3. python import_to_mongo.py
"""

import openpyxl
import datetime
from pymongo import MongoClient, GEOSPHERE
from pymongo.errors import BulkWriteError

# ── CONFIGURACIÓ ──────────────────────────────────────────────────────────────

MONGO_URI  = "mongodb+srv://pautorrejon_db_user:tQDV09cNAMgoELai@dammsmarttruck.k9q2nbf.mongodb.net/damm_hackathon?retryWrites=true&w=majority&appName=DammSmartTruck"
DB_NAME    = "damm_hackathon"
EXCEL_DIR  = r"C:\Users\torre\Documents\UNI\Hackaton\Hackaton_files\Datos"

# Coordenades GPS reals dels clients de Granollers
# (obtingudes via geocodificació de les adreces dels Excels)
GPS_GRANOLLERS = {
    "RESTAURANTE GRAN MURALLA"  : (41.6087, 2.2940),
    "A VOCADOS"                 : (41.6063, 2.2862),
    "CAN BIN RESTAURANTE"       : (41.6092, 2.2931),
    "CAFE DE LA CORONA"         : (41.6071, 2.2877),
    "BK GRANOLLERS"             : (41.6021, 2.2810),
    "LAPONIA CAFE"              : (41.6055, 2.2869),
    "Comercial  Bebidas Burj SL": (41.5998, 2.2950),
    "BAR LA CANTINA"            : (41.6078, 2.2925),
    "CAFE SANT ROC"             : (41.6068, 2.2890),
    "VIENA GRANOLLERS (ANSELM CLAVE) 10": (41.6075, 2.2883),
}

# Zones de vianants de Granollers (super-nodes d'aparcament)
PEDESTRIAN_ZONES = [
    {
        "zone_id"     : "PZ_CORONA",
        "nom"         : "Zona vianants Plaça Corona",
        "parking_lat" : 41.6075,
        "parking_lon" : 2.2880,
        "radi_m"      : 150,
        "client_noms" : ["CAFE DE LA CORONA", "LAPONIA CAFE", "A VOCADOS"],
        "minuts_a_peu": {"CAFE DE LA CORONA": 2, "LAPONIA CAFE": 4, "A VOCADOS": 5}
    },
    {
        "zone_id"     : "PZ_SANT_ROC",
        "nom"         : "Zona vianants Carrer Sant Roc",
        "parking_lat" : 41.6070,
        "parking_lon" : 2.2895,
        "radi_m"      : 100,
        "client_noms" : ["CAFE SANT ROC"],
        "minuts_a_peu": {"CAFE SANT ROC": 3}
    }
]


def connect_mongo():
    """Connecta a MongoDB Atlas i retorna la base de dades."""
    try:
        client = MongoClient(MONGO_URI, serverSelectionTimeoutMS=5000)
        client.admin.command("ping")
        print("✓ Connexió a MongoDB Atlas establerta")
        return client[DB_NAME]
    except Exception as e:
        print(f"✗ Error de connexió: {e}")
        print("  Comprova la contrasenya al MONGO_URI")
        raise


def import_products(db):
    """
    Importa productes de ZM040.XLSX.
    Filtra únicament les files amb unitat CAJ i dimensions > 0.
    """
    print("\n[1/4] Important productes (ZM040.XLSX)...")
    wb = openpyxl.load_workbook(f"{EXCEL_DIR}/ZM040.XLSX", data_only=True)
    ws = wb.active

    products = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # capçalera

        codi     = str(row[0]).strip() if row[0] else None
        unitat   = str(row[2]).strip() if row[2] else None
        nom      = str(row[4]).strip() if row[4] else None
        llarg    = row[6]   # Longitud cm
        ample    = row[8]   # Ancho cm
        alt      = row[10]  # Altura cm
        volum    = row[13]  # Volumen L
        pes_brut = row[16]  # Peso bruto kg

        # Només importem caixes (CAJ) amb dimensions vàlides
        if unitat != "CAJ":
            continue
        if not all([llarg, ample, alt]) or llarg == 0:
            continue
        if not codi or codi == "None":
            continue

        products[codi] = {
            "codi"      : codi,
            "nom"       : nom or codi,
            "unitat"    : "CAJ",
            "llarg_cm"  : float(llarg),
            "ample_cm"  : float(ample),
            "alt_cm"    : float(alt),
            "volum_l"   : float(volum) if volum else 0,
            "pes_kg"    : float(pes_brut) if pes_brut else 0,
            "actualitzat": datetime.datetime.utcnow()
        }

    col = db["products"]
    col.drop()  # neteja per reimportació neta
    if products:
        col.insert_many(list(products.values()))
        col.create_index("codi", unique=True)
        print(f"  ✓ {len(products)} productes importats")
    return products


def import_time_windows(db):
    """
    Importa horaris d'entrega de Horarios Entrega.XLSX.
    Retorna dict {client_nom: {dia_setmana: (open_min, close_min)}}
    """
    print("\n[2/4] Important horaris (Horarios Entrega.XLSX)...")
    wb = openpyxl.load_workbook(
        f"{EXCEL_DIR}/Horarios Entrega.XLSX", data_only=True
    )
    ws = wb.active

    horaris = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue

        deudor  = str(row[0]).strip() if row[0] else None
        nom     = str(row[6]).strip() if row[6] else None
        dia     = row[4]   # 1=Dilluns ... 7=Diumenge
        hora_ob = row[10]  # datetime.time
        hora_cl = row[11]

        if not nom or not dia:
            continue

        def time_to_min(t):
            if not t or not isinstance(t, datetime.time):
                return None
            return t.hour * 60 + t.minute

        ob_min = time_to_min(hora_ob)
        cl_min = time_to_min(hora_cl)

        # Si l'horari de tancament és 0 significa tancat
        if cl_min == 0:
            continue

        if nom not in horaris:
            horaris[nom] = {"deudor": deudor, "dies": {}}

        horaris[nom]["dies"][str(dia)] = {
            "open_min" : ob_min if ob_min is not None else 0,
            "close_min": cl_min if cl_min is not None else 23 * 60 + 59
        }

    print(f"  ✓ Horaris de {len(horaris)} clients carregats")
    return horaris


def import_clients_and_orders(db, horaris, products):
    """
    Importa clients i comandes de Hackaton.xlsx.
    Filtra únicament clients de Granollers.
    Agrupa les línies de comanda per client.
    """
    print("\n[3/4] Important clients i comandes (Hackaton.xlsx)...")
    wb = openpyxl.load_workbook(
        f"{EXCEL_DIR}/Hackaton.xlsx", data_only=True
    )
    ws = wb.active

    clients_dict = {}
    orders_dict  = {}

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue

        poblacio = str(row[15]).strip().upper() if row[15] else ""
        if "GRANOLLERS" not in poblacio:
            continue

        # Dades del client
        repartidor  = str(row[3]).strip() if row[3] else None
        client_nom  = str(row[11]).strip() if row[11] else None
        client_nom2 = str(row[12]).strip() if row[12] else ""
        carrer      = str(row[13]).strip() if row[13] else ""
        cp          = str(row[14]).strip() if row[14] else ""
        client_id   = str(row[4]).strip() if row[4] else None
        entrega_id  = str(row[5]).strip() if row[5] else None

        # Dades del producte
        material    = str(row[6]).strip() if row[6] else None
        denominacio = str(row[7]).strip() if row[7] else None
        quantitat   = row[8] if row[8] else 0
        unitat_vd   = str(row[9]).strip() if row[9] else ""

        if not client_nom or not material:
            continue

        # Coordenades GPS
        coords = GPS_GRANOLLERS.get(client_nom, (41.6067, 2.2894))  # centre Granollers per defecte

        # Detectar zona de vianants
        zona_vianants = None
        for pz in PEDESTRIAN_ZONES:
            if client_nom in pz["client_noms"]:
                zona_vianants = pz["zone_id"]
                break

        # Construir document client (GeoJSON per a índex geoespacial)
        if client_nom not in clients_dict:
            # Obtenim finestres horàries (default: 8:00-22:00)
            horari_client = horaris.get(client_nom, {})
            tw_default = {"open_min": 8 * 60, "close_min": 22 * 60}
            tw_dilluns = horari_client.get("dies", {}).get("1", tw_default)

            clients_dict[client_nom] = {
                "client_id"       : client_id,
                "nom"             : client_nom,
                "nom2"            : client_nom2,
                "carrer"          : carrer,
                "cp"              : cp,
                "poblacio"        : "Granollers",
                "location"        : {          # GeoJSON per a consultes geoespacials
                    "type"        : "Point",
                    "coordinates" : [coords[1], coords[0]]  # [lon, lat]
                },
                "lat"             : coords[0],
                "lon"             : coords[1],
                "zona_vianants_id": zona_vianants,
                "open_min"        : tw_dilluns["open_min"],
                "close_min"       : tw_dilluns["close_min"],
                "service_min"     : 10,        # temps estimat de descàrrega
                "actualitzat"     : datetime.datetime.utcnow()
            }

        # Construir comanda
        order_key = f"{client_nom}_{entrega_id}"
        if order_key not in orders_dict:
            orders_dict[order_key] = {
                "entrega_id" : entrega_id,
                "client_nom" : client_nom,
                "client_id"  : client_id,
                "repartidor" : repartidor,
                "poblacio"   : "Granollers",
                "linies"     : [],
                "kg_total"   : 0.0,
                "vol_total_l": 0.0,
                "data"       : datetime.datetime.utcnow()
            }

        # Dimensió del producte
        prod_info = products.get(material, {})
        pes_linia = prod_info.get("pes_kg", 0) * float(quantitat)
        vol_linia = prod_info.get("volum_l", 0) * float(quantitat)

        orders_dict[order_key]["linies"].append({
            "material"   : material,
            "nom"        : denominacio,
            "quantitat"  : float(quantitat),
            "unitat"     : unitat_vd,
            "pes_kg"     : pes_linia,
            "vol_l"      : vol_linia,
            "llarg_cm"   : prod_info.get("llarg_cm", 0),
            "ample_cm"   : prod_info.get("ample_cm", 0),
            "alt_cm"     : prod_info.get("alt_cm", 0),
        })
        orders_dict[order_key]["kg_total"]   += pes_linia
        orders_dict[order_key]["vol_total_l"] += vol_linia

    # Importar clients
    col_clients = db["clients"]
    col_clients.drop()
    if clients_dict:
        col_clients.insert_many(list(clients_dict.values()))
        col_clients.create_index([("location", GEOSPHERE)])  # índex geoespacial
        col_clients.create_index("nom", unique=True)
        print(f"  ✓ {len(clients_dict)} clients de Granollers importats")

    # Importar comandes
    col_orders = db["orders"]
    col_orders.drop()
    if orders_dict:
        col_orders.insert_many(list(orders_dict.values()))
        col_orders.create_index("client_nom")
        col_orders.create_index("entrega_id")
        print(f"  ✓ {len(orders_dict)} comandes importades")

    return clients_dict, orders_dict


def import_pedestrian_zones(db):
    """Importa les zones de vianants de Granollers."""
    print("\n[4/4] Important zones de vianants...")
    col = db["pedestrian_zones"]
    col.drop()
    col.insert_many(PEDESTRIAN_ZONES)
    col.create_index([("parking_lat", 1), ("parking_lon", 1)])
    print(f"  ✓ {len(PEDESTRIAN_ZONES)} zones de vianants importades")


def verify_import(db):
    """Verifica que tot s'ha importat correctament."""
    print("\n" + "=" * 55)
    print("  VERIFICACIÓ FINAL")
    print("=" * 55)
    coleccions = ["products", "clients", "orders", "pedestrian_zones"]
    for col_name in coleccions:
        n = db[col_name].count_documents({})
        print(f"  {col_name:<20} → {n:>4} documents")

    # Exemple de client amb la seva comanda
    client = db["clients"].find_one({"poblacio": "Granollers"})
    if client:
        order = db["orders"].find_one({"client_nom": client["nom"]})
        print(f"\n  Exemple client: {client['nom']}")
        print(f"    Adreça  : {client['carrer']}")
        print(f"    Horari  : {client['open_min']//60:02d}:{client['open_min']%60:02d}"
              f" – {client['close_min']//60:02d}:{client['close_min']%60:02d}")
        if client.get("zona_vianants_id"):
            print(f"    Vianants: {client['zona_vianants_id']}")
        if order:
            print(f"    Comanda : {len(order['linies'])} línies | "
                  f"{order['kg_total']:.1f} kg")
    print("=" * 55)
    print("  Base de dades llesta. Pots arrencar el backend!")
    print("=" * 55)


if __name__ == "__main__":
    print("=" * 55)
    print("  DAMM SMART TRUCK · Import MongoDB Atlas")
    print("  Granollers · Interhack BCN 2026")
    print("=" * 55)

    try:
        db       = connect_mongo()
        products = import_products(db)
        horaris  = import_time_windows(db)
        import_clients_and_orders(db, horaris, products)
        import_pedestrian_zones(db)
        verify_import(db)
    except Exception as e:
        print(f"\n✗ Error durant la importació: {e}")
        raise
